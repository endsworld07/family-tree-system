from __future__ import annotations

import json
import re
import textwrap
from io import BytesIO
from datetime import datetime
from pathlib import Path
from typing import Dict
from uuid import uuid4

from flask import Flask, redirect, render_template, request, send_file, url_for

from connection_engine import ConnectionEngine
from data_loader import DataLoader
from layout_engine import LayoutEngine
from models import Person
from relationship_engine import RelationshipEngine
from svg_renderer import SvgRenderer

app = Flask(__name__)

DATA_FILE = Path("family.json")
SAVES_FILE = Path("saved_families.json")


# ---------------------------------------------------------
# 讀取資料
# ---------------------------------------------------------

def load_people() -> tuple[Dict[str, Person], str]:
    if not DATA_FILE.exists():
        return {}, ""

    loader = DataLoader(str(DATA_FILE))
    data = loader.load()

    return data["people"], data["applicant"]


# ---------------------------------------------------------
# 儲存資料
# ---------------------------------------------------------

def people_payload(
    people: Dict[str, Person],
    applicant_id: str,
) -> dict:

    payload = {
        "applicant": applicant_id,
        "people": [],
    }

    for person in people.values():

        payload["people"].append(
            {
                "id": person.id,
                "name": person.name,
                "label": person.label,
                "father": person.father,
                "mother": person.mother,
                "spouses": list(person.spouses),
                "is_exhumation": person.is_exhumation,
                "non_exhumation_note": person.non_exhumation_note,
            }
        )

    return payload


def write_payload(filename: Path, payload: dict) -> None:
    filename.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def save_people(
    people: Dict[str, Person],
    applicant_id: str,
) -> None:
    write_payload(DATA_FILE, people_payload(people, applicant_id))


def load_saved_families() -> list[dict]:
    if not SAVES_FILE.exists():
        return []
    try:
        data = json.loads(SAVES_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    return data.get("saves", [])


def save_named_family(name: str, people: Dict[str, Person], applicant_id: str) -> None:
    saves = load_saved_families()
    record = {
        "id": uuid4().hex,
        "name": name,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "payload": people_payload(people, applicant_id),
    }
    for index, saved in enumerate(saves):
        if saved.get("name") == name:
            record["id"] = saved.get("id", record["id"])
            saves[index] = record
            break
    else:
        saves.append(record)
    write_payload(SAVES_FILE, {"saves": saves})


# ---------------------------------------------------------
# 產生 SVG
# ---------------------------------------------------------

def build_svg_with_layout(
    people: Dict[str, Person],
    applicant_id: str,
) -> tuple[str, object | None]:

    if (
        not people
        or not applicant_id
        or applicant_id not in people
    ):
        return "", None

    relationship_engine = RelationshipEngine(people)
    relationships = relationship_engine.build()

    layout_engine = LayoutEngine(
        people,
        relationships,
        applicant_id,
    )

    layout_result = layout_engine.build()

    connection_engine = ConnectionEngine(
        people,
        relationships,
        layout_result,
    )

    connection_result = connection_engine.build()

    renderer = SvgRenderer()

    svg_result = renderer.render(
        people,
        layout_result,
        connection_result,
    )

    return svg_result.svg, layout_result


def build_svg(
    people: Dict[str, Person],
    applicant_id: str,
) -> str:
    return build_svg_with_layout(people, applicant_id)[0]


def people_not_in_svg(people: Dict[str, Person], layout_result) -> set[str]:
    """Return people that the current relationship layout cannot place on the graph."""
    if layout_result is None:
        return set(people)
    return set(people) - set(layout_result.positions)


# ---------------------------------------------------------
# 下一個人物 ID
# ---------------------------------------------------------

def next_person_id(
    people: Dict[str, Person],
) -> str:

    max_number = 0

    for pid in people.keys():

        if pid.startswith("P") and pid[1:].isdigit():

            max_number = max(
                max_number,
                int(pid[1:]),
            )

    return f"P{max_number + 1:03d}"


# ---------------------------------------------------------
# 依姓名取得人物
# 找不到就建立
# ---------------------------------------------------------

def get_or_create_person(
    people: Dict[str, Person],
    name: str,
) -> str | None:

    name = name.strip()

    if not name:
        return None

    # 已存在
    for pid, person in people.items():

        if person.name == name:
            return pid

    # 建立新人物
    new_id = next_person_id(people)

    people[new_id] = Person(
        id=new_id,
        name=name,
        label="",
        father=None,
        mother=None,
        spouses=[],
        is_exhumation=False,
        non_exhumation_note="",
    )

    return new_id


def spouse_names_from_form() -> list[str]:
    """Read one or more spouse names from the form, preserving entry order."""
    raw = request.form.get("spouses", request.form.get("spouse", ""))
    names: list[str] = []
    for name in re.split(r"[,，、\n]", raw):
        name = name.strip()
        if name and name not in names:
            names.append(name)
    return names


def update_spouses(people: Dict[str, Person], person_id: str, spouse_ids: list[str]) -> None:
    """Replace a person's spouse list and keep every relationship reciprocal."""
    for other in people.values():
        if person_id in other.spouses:
            other.spouses.remove(person_id)

    person = people[person_id]
    person.spouses = []
    for spouse_id in spouse_ids:
        if spouse_id == person_id or spouse_id not in people:
            continue
        if spouse_id not in person.spouses:
            person.spouses.append(spouse_id)
        if person_id not in people[spouse_id].spouses:
            people[spouse_id].spouses.append(person_id)


def exhumation_count(people: Dict[str, Person]) -> int:
    return sum(person.is_exhumation for person in people.values())


def svg_png(svg: str, people: Dict[str, Person], layout_result) -> BytesIO:
    """Convert the graph to PNG and paint Chinese labels with the bundled font."""
    import cairosvg
    from PIL import Image, ImageChops, ImageDraw, ImageFont

    raw_image = BytesIO()
    scale = 2
    cairosvg.svg2png(bytestring=svg.encode("utf-8"), write_to=raw_image, scale=scale)
    raw_image.seek(0)
    rendered = Image.open(raw_image).convert("RGBA")
    image = Image.new("RGBA", rendered.size, "white")
    image.alpha_composite(rendered)

    font_path = Path(__file__).with_name("fonts") / "NotoSansCJKtc-Regular.otf"
    if not font_path.exists():
        raise RuntimeError("找不到繁中字型檔 fonts/NotoSansCJKtc-Regular.otf")
    name_font = ImageFont.truetype(str(font_path), 14 * scale)
    label_font = ImageFont.truetype(str(font_path), 13 * scale)
    note_font = ImageFont.truetype(str(font_path), 9 * scale)
    draw = ImageDraw.Draw(image)

    positions = layout_result.positions
    min_column = min(position.column for position in positions.values())
    min_row = min(position.row for position in positions.values())
    origin_x = 60 - min_column * 180
    origin_y = 60 - min_row * 75
    for person_id, position in positions.items():
        person = people.get(person_id)
        if person is None:
            continue
        x = int((origin_x + position.column * 180 - 60) * scale)
        y = int((origin_y + position.row * 75 - 30) * scale)
        width, height = 120 * scale, 60 * scale
        is_applicant = person.label.strip() == "申請人"
        fill = "#fff3a3" if is_applicant else ("#ffffff" if person.is_exhumation else "#eeeeee")
        outline = "#c78300" if is_applicant else ("#000000" if person.is_exhumation else "#999999")
        text_color = "#000000" if is_applicant or person.is_exhumation else "#777777"
        draw.rectangle((x, y, x + width, y + height), fill=fill, outline=outline, width=scale)
        center_x = x + width // 2
        note = person.non_exhumation_note.strip() if not person.is_exhumation else ""
        if note:
            draw.text((center_x, y + 18 * scale), person.name, font=name_font, fill=text_color, anchor="mm")
            draw.text((center_x, y + 35 * scale), person.label or "", font=label_font, fill=text_color, anchor="mm")
            draw.text((center_x, y + 51 * scale), note, font=note_font, fill=text_color, anchor="mm")
        else:
            draw.text((center_x, y + 24 * scale), person.name, font=name_font, fill=text_color, anchor="mm")
            draw.text((center_x, y + 42 * scale), person.label or "", font=label_font, fill=text_color, anchor="mm")

    # Remove renderer padding so the relationship graph itself can fill the A4 graph area.
    rgb_image = image.convert("RGB")
    white_background = Image.new("RGB", rgb_image.size, "white")
    content_box = ImageChops.difference(rgb_image, white_background).getbbox()
    if content_box:
        padding = 12 * scale
        left, top, right, bottom = content_box
        image = image.crop((
            max(0, left - padding),
            max(0, top - padding),
            min(image.width, right + padding),
            min(image.height, bottom + padding),
        ))

    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return output


def export_pdf(graph_png: BytesIO | None, count: int, recipient: str) -> BytesIO:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas

    font_path = Path(__file__).with_name("fonts") / "NotoSansTC-Variable.ttf"
    if not font_path.exists():
        raise RuntimeError("找不到繁中字型檔 fonts/NotoSansTC-Variable.ttf")
    font = "NotoSansTC"
    pdfmetrics.registerFont(TTFont(font, str(font_path)))
    page_width, page_height = A4
    margin = 28.35  # 1 cm
    output = BytesIO()
    document = canvas.Canvas(output, pagesize=A4)

    document.setFont(font, 18)
    title = "親屬關係表"
    title_gap = 28.35  # 1 cm between each character
    title_width = sum(pdfmetrics.stringWidth(character, font, 18) for character in title)
    title_width += title_gap * (len(title) - 1)
    title_x = (page_width - title_width) / 2
    for character in title:
        document.drawString(title_x, page_height - margin - 18, character)
        title_x += pdfmetrics.stringWidth(character, font, 18) + title_gap
    if graph_png:
        image = ImageReader(graph_png)
        graph_bottom = 215
        graph_top = page_height - margin - 42
        document.drawImage(
            image,
            margin,
            graph_bottom,
            width=page_width - margin * 2,
            height=graph_top - graph_bottom,
            preserveAspectRatio=True,
            anchor="c",
        )
    statement_lines = [
        f"以上共{count}位被申請起掘，其確係本人祖先（關係或稱謂如上表）無訛，且均由本人祭拜，",
        "故由本人申請辦理遷葬事宜，特立切結，如有虛偽造假及相關法律糾紛，概由本人負責，與貴所無涉。",
    ]

    def draw_justified_line(line: str, baseline_y: float, underline_count: bool = False) -> None:
        """Draw a line whose first and last characters align to the 1 cm page margins."""
        character_widths = [pdfmetrics.stringWidth(character, font, 10) for character in line]
        available_width = page_width - margin * 2
        spacing = (available_width - sum(character_widths)) / max(1, len(line) - 1)
        count_start = line.find(str(count)) if underline_count else -1
        count_indexes = set(range(count_start, count_start + len(str(count))))
        x = margin
        for index, (character, character_width) in enumerate(zip(line, character_widths)):
            document.drawString(x, baseline_y, character)
            if index in count_indexes:
                document.line(x, baseline_y - 1.5, x + character_width, baseline_y - 1.5)
            x += character_width + spacing

    y = 192
    document.setFont(font, 10)
    for index, line in enumerate(statement_lines):
        draw_justified_line(line, y, underline_count=index == 0)
        y -= 15
    y -= 4
    document.drawString(margin, y, "此致")
    y -= 18
    document.drawString(margin, y, recipient or "____________________________")
    y -= 31
    document.drawCentredString(page_width / 2, y, "具結人：___________________（簽章）")
    y -= 25
    document.drawCentredString(page_width / 2, y, "中華民國 ____ 年 ____ 月 ____ 日")
    document.save()
    output.seek(0)
    return output


def export_xlsx(graph_png: BytesIO | None, count: int, recipient: str) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Font
    from openpyxl.worksheet.page import PageMargins

    workbook = Workbook()
    chart_sheet = workbook.active
    chart_sheet.title = "親屬關係表"
    chart_sheet.page_setup.paperSize = chart_sheet.PAPERSIZE_A4
    chart_sheet.page_setup.orientation = "landscape"
    chart_sheet.page_setup.fitToWidth = 1
    chart_sheet.page_setup.fitToHeight = 1
    chart_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    chart_sheet.page_margins = PageMargins(left=0.3937, right=0.3937, top=0.3937, bottom=0.3937, header=0, footer=0)
    chart_sheet.print_options.horizontalCentered = True
    chart_sheet.print_options.verticalCentered = True
    chart_sheet.sheet_view.showGridLines = False
    chart_sheet.merge_cells("A1:H2")
    title = chart_sheet["A1"]
    title.value = "親屬關係表"
    title.font = Font(name="Microsoft JhengHei", size=20, bold=True)
    title.alignment = Alignment(horizontal="center", vertical="center")
    if graph_png:
        image = Image(graph_png)
        original_width, original_height = image.width, image.height
        scale = min(820 / original_width, 330 / original_height)
        image.width = original_width * scale
        image.height = original_height * scale
        chart_sheet.add_image(image, "A3")
    for column in "ABCDEFGH":
        chart_sheet.column_dimensions[column].width = 16
    for row in range(1, 39):
        chart_sheet.row_dimensions[row].height = 18
    chart_sheet.row_dimensions[1].height = 28
    chart_sheet.row_dimensions[2].height = 20
    chart_sheet.merge_cells("A23:H27")
    chart_sheet["A23"] = (
        f"以上共 {count} 位被申請起掘，其確係本人祖先（關係或稱謂如上表）無訛，"
        "且均由本人祭拜，故由本人申請辦理遷葬事宜，特立切結，如有虛偽造假及相關法律糾紛，"
        "概由本人負責，與貴所無涉。"
    )
    chart_sheet["A23"].alignment = Alignment(wrap_text=True, vertical="top")
    chart_sheet["A23"].font = Font(name="Microsoft JhengHei", size=10)
    chart_sheet.merge_cells("A29:H29")
    chart_sheet["A29"] = "此致"
    chart_sheet.merge_cells("A30:H30")
    chart_sheet["A30"] = recipient or "____________________________"
    chart_sheet.merge_cells("A33:H33")
    chart_sheet["A33"] = "具結人：___________________（簽章）"
    chart_sheet.merge_cells("A36:H36")
    chart_sheet["A36"] = "中華民國 ____ 年 ____ 月 ____ 日"
    chart_sheet.print_area = "A1:H38"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------
# 首頁
# ---------------------------------------------------------

@app.get("/")
def index():

    people, applicant_id = load_people()

    svg, layout_result = build_svg_with_layout(people, applicant_id)

    return render_template(
        "index.html",
        people=people,
        applicant_id=applicant_id,
        svg=svg,
        hidden_person_ids=people_not_in_svg(people, layout_result),
        exhumation_count=exhumation_count(people),
        saved_families=load_saved_families(),
    )



@app.post("/person/add")
def person_add():

    people, applicant_id = load_people()

    name = request.form.get("name", "").strip()
    label = request.form.get("label", "").strip()

    father_name = request.form.get("father", "").strip()
    mother_name = request.form.get("mother", "").strip()
    spouse_names = spouse_names_from_form()
    is_exhumation = request.form.get("is_exhumation") == "on"
    non_exhumation_note = request.form.get("non_exhumation_note", "").strip()

    if not name:
        return redirect(url_for("index"))

    # 已存在的人物直接返回
    for person in people.values():
        if person.name == name:
            return redirect(url_for("index"))

    # 建立或取得父母、配偶
    father_id = get_or_create_person(
        people,
        father_name,
    )

    mother_id = get_or_create_person(
        people,
        mother_name,
    )

    spouse_ids = [
        spouse_id
        for spouse_id in (get_or_create_person(people, spouse_name) for spouse_name in spouse_names)
        if spouse_id
    ]

    person_id = next_person_id(people)

    people[person_id] = Person(
        id=person_id,
        name=name,
        label=label,
        father=father_id,
        mother=mother_id,
        spouses=[],
        is_exhumation=is_exhumation,
        non_exhumation_note="" if is_exhumation else non_exhumation_note,
    )

    update_spouses(people, person_id, spouse_ids)

    if not applicant_id:
        applicant_id = person_id

    save_people(people, applicant_id)

    print(">>> person_edit()")

    return redirect(url_for("index"))

# ---------------------------------------------------------
# 編輯頁面
# ---------------------------------------------------------

@app.get("/person/edit/<person_id>")
def person_edit(person_id: str):

    people, applicant_id = load_people()

    if person_id not in people:
        return redirect(url_for("index"))

    svg, layout_result = build_svg_with_layout(people, applicant_id)
    print(">>> person_edit()")
    return render_template(
        "index.html",
        people=people,
        applicant_id=applicant_id,
        svg=svg,
        hidden_person_ids=people_not_in_svg(people, layout_result),
        editing=people[person_id],
        editing_spouse_names="、".join(
            people[spouse_id].name
            for spouse_id in people[person_id].spouses
            if spouse_id in people
        ),
        exhumation_count=exhumation_count(people),
        saved_families=load_saved_families(),
    )


# ---------------------------------------------------------
# 更新人物
# ---------------------------------------------------------

@app.post("/person/update/<person_id>")
def person_update(person_id: str):

    people, applicant_id = load_people()

    if person_id not in people:
        return redirect(url_for("index"))

    person = people[person_id]

    name = request.form.get("name", "").strip()
    label = request.form.get("label", "").strip()

    father_name = request.form.get("father", "").strip()
    mother_name = request.form.get("mother", "").strip()
    spouse_names = spouse_names_from_form()
    is_exhumation = request.form.get("is_exhumation") == "on"
    non_exhumation_note = request.form.get("non_exhumation_note", "").strip()

    if name:
        person.name = name

    person.label = label

    father_id = get_or_create_person(
        people,
        father_name,
    )

    mother_id = get_or_create_person(
        people,
        mother_name,
    )

    spouse_ids = [
        spouse_id
        for spouse_id in (get_or_create_person(people, spouse_name) for spouse_name in spouse_names)
        if spouse_id
    ]

    # 避免自己指向自己
    if father_id == person_id:
        father_id = None

    if mother_id == person_id:
        mother_id = None

    person.father = father_id
    person.mother = mother_id
    person.is_exhumation = is_exhumation
    person.non_exhumation_note = "" if is_exhumation else non_exhumation_note
    update_spouses(people, person_id, spouse_ids)

    save_people(
        people,
        applicant_id,
    )

    return redirect(
        url_for("index")
    )

# ---------------------------------------------------------
# 刪除人物
# ---------------------------------------------------------

@app.post("/person/delete/<person_id>")
def person_delete(person_id: str):

    people, applicant_id = load_people()

    if person_id not in people:
        return redirect(url_for("index"))

    # 清除所有人物對此人物的關聯
    for other in people.values():

        if other.father == person_id:
            other.father = None

        if other.mother == person_id:
            other.mother = None

        if person_id in other.spouses:
            other.spouses.remove(person_id)

    # 刪除人物
    del people[person_id]

    # 若刪除的是申請人
    if applicant_id == person_id:

        if people:
            applicant_id = next(iter(people.keys()))
        else:
            applicant_id = ""

    save_people(
        people,
        applicant_id,
    )

    return redirect(
        url_for("index")
    )


# ---------------------------------------------------------
# 設定申請人
# ---------------------------------------------------------

@app.post("/set-applicant/<person_id>")
def set_applicant(person_id: str):

    people, applicant_id = load_people()

    if person_id not in people:
        return redirect(url_for("index"))

    applicant_id = person_id

    save_people(
        people,
        applicant_id,
    )

    return redirect(
        url_for("index")
    )


# ---------------------------------------------------------
# 重新產生 SVG
# ---------------------------------------------------------

@app.post("/generate")
def generate():

    people, applicant_id = load_people()

    svg, layout_result = build_svg_with_layout(people, applicant_id)

    return render_template(
        "index.html",
        people=people,
        applicant_id=applicant_id,
        svg=svg,
        hidden_person_ids=people_not_in_svg(people, layout_result),
        exhumation_count=exhumation_count(people),
        saved_families=load_saved_families(),
    )


# ---------------------------------------------------------
# 命名儲存與載入
# ---------------------------------------------------------

@app.post("/save-as")
def save_as():
    people, applicant_id = load_people()
    name = request.form.get("save_name", "").strip()
    if name:
        save_named_family(name, people, applicant_id)
    return redirect(url_for("index"))


@app.post("/load-save/<save_id>")
def load_save(save_id: str):
    for saved in load_saved_families():
        if saved.get("id") == save_id and isinstance(saved.get("payload"), dict):
            write_payload(DATA_FILE, saved["payload"])
            break
    return redirect(url_for("index"))


@app.post("/delete-save/<save_id>")
def delete_save(save_id: str):
    saves = [saved for saved in load_saved_families() if saved.get("id") != save_id]
    write_payload(SAVES_FILE, {"saves": saves})
    return redirect(url_for("index"))


@app.post("/clear-people")
def clear_people():
    """Clear only the current working family; named saves remain available."""
    save_people({}, "")
    return redirect(url_for("index"))


@app.post("/export")
def export_document():
    people, applicant_id = load_people()
    recipient = request.form.get("recipient", "").strip()
    export_format = request.form.get("export_format", "pdf")
    svg, layout_result = build_svg_with_layout(people, applicant_id)
    graph_png = svg_png(svg, people, layout_result) if svg and layout_result else None
    count = exhumation_count(people)

    if export_format == "xlsx":
        return send_file(
            export_xlsx(graph_png, count, recipient),
            as_attachment=True,
            download_name="親屬關係表.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return send_file(
        export_pdf(graph_png, count, recipient),
        as_attachment=True,
        download_name="親屬關係表.pdf",
        mimetype="application/pdf",
    )


# ---------------------------------------------------------
# 主程式
# ---------------------------------------------------------

if __name__ == "__main__":
    app.run(
        debug=True,
    )
